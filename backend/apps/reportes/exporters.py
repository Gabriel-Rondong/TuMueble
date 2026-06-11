"""
Generadores de reportes en Excel (openpyxl) y PDF (reportlab).
"""
import io
from django.http import HttpResponse
from django.utils import timezone


def exportar_excel(data: list[dict], titulo: str, columnas: list[str]) -> HttpResponse:
    """Genera un archivo Excel con los datos proporcionados."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl no instalado", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # Excel max 31 chars

    # Encabezado empresa
    ws.merge_cells("A1:E1")
    ws["A1"] = f"TuMueble ERP — {titulo}"
    ws["A1"].font = Font(bold=True, size=14, color="C9963A")
    ws["A1"].fill = PatternFill("solid", fgColor="1A1D24")

    ws["A2"] = f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="7A8099", size=9)

    # Headers
    header_fill = PatternFill("solid", fgColor="21252F")
    header_font = Font(bold=True, color="E8EAF0", size=10)
    thin = Side(style="thin", color="2D3240")
    border = Border(bottom=thin)

    for col_i, col_name in enumerate(columnas, 1):
        cell = ws.cell(row=4, column=col_i, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Datos
    for row_i, row in enumerate(data, 5):
        for col_i, key in enumerate(columnas, 1):
            val = row.get(key, row.get(list(row.keys())[col_i - 1] if col_i <= len(row) else "", ""))
            cell = ws.cell(row=row_i, column=col_i, value=val)
            if row_i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="1A1D24")

    # Ajustar anchos
    for col_i in range(1, len(columnas) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = titulo.replace(" ", "_").lower()
    response["Content-Disposition"] = f'attachment; filename="{filename}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    return response


def exportar_pdf_simple(data: list[dict], titulo: str, columnas: list[str]) -> HttpResponse:
    """Genera un PDF simple con tabla de datos usando reportlab."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        return HttpResponse("reportlab no instalado", status=500)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    COLOR_GOLD = colors.HexColor("#C9963A")
    COLOR_DARK = colors.HexColor("#1A1D24")
    COLOR_TEXT = colors.HexColor("#E8EAF0")

    styles = getSampleStyleSheet()
    elements = []

    # Título
    title_style = ParagraphStyle("title", parent=styles["Title"],
                                  textColor=COLOR_GOLD, fontSize=16, spaceAfter=4)
    elements.append(Paragraph(f"TuMueble ERP — {titulo}", title_style))

    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                                textColor=colors.HexColor("#7A8099"), fontSize=8, spaceAfter=12)
    elements.append(Paragraph(f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}", sub_style))
    elements.append(Spacer(1, 0.3*cm))

    # Tabla
    table_data = [columnas]
    for row in data:
        vals = list(row.values())
        table_data.append([str(v) if v is not None else "—" for v in vals])

    col_width = (landscape(A4)[0] - 3*cm) / max(len(columnas), 1)
    t = Table(table_data, colWidths=[col_width] * len(columnas), repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_GOLD),
        ("TEXTCOLOR", (0, 0), (-1, 0), COLOR_DARK),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [COLOR_DARK, colors.HexColor("#21252F")]),
        ("TEXTCOLOR", (0, 1), (-1, -1), COLOR_TEXT),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#2D3240")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    filename = titulo.replace(" ", "_").lower()
    response["Content-Disposition"] = f'attachment; filename="{filename}_{timezone.now().strftime("%Y%m%d")}.pdf"'
    return response